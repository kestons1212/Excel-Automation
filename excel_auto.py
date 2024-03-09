import openpyxl
from openpyxl import Workbook, load_workbook
import string
import csv
import re
import os



# files for comparison

file1 = 'Cars.xlsx'
file2 = 'Cars1.xlsx'

wb1 = load_workbook(file1)
wb2 = load_workbook(file2)


# Files for keeping a count of the data.

file3 = 'Cars.csv'
file4 = 'Cars1.csv'

data1 = []
data2 = []

with open(file3, newline='') as csvfile1:
    stuff1 = csv.reader(csvfile1, delimiter=',', quotechar='|')
    data1.extend(stuff1)

with open(file4, newline='') as csvfile2:
    stuff2 = csv.reader(csvfile2, delimiter=',', quotechar='|')
    data2.extend(stuff2)


# Excel work books.

ws1 = wb1.active
ws2 = wb2.active


# number and letter data base for row/coloumn indexing.

num = [str(x + 1) for x in range(0,len(data2))]
letters = list(string.ascii_uppercase)

count = 0

I = ''

for i in range(0, len(data1)):
    for j in range(0 ,len(data2)):

        wordA = ws1[letters[0] + num[i]].value
        wordB = ws2[letters[0] + num[j ]].value

        wordA_comp = re.sub('[\W_]+', '',wordA.lower() )
        wordB_comp = re.sub('[\W_]+', '',wordB.lower() )
       
        if wordA_comp == wordB_comp:
            
            I += str(ws2[letters[3] + num[j]].value) + ','
            
            ws1[letters[5] + num[i]].value = wordB = ws2[letters[0] + num[j ]].value
            ws1[letters[6] + num[i]].value = I

wb1.save(file1)
wb2.save(file2)



os.system(file1)
# os.system(file2)
